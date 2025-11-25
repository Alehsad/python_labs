import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    csv_file = Path(csv_path)
    xlsx_file = Path(xlsx_path)

    if csv_file.suffix.lower() != ".csv" or xlsx_file.suffix.lower() != ".xlsx":
        raise ValueError("Ожидаются файлы .csv (вход) и .xlsx (выход)")

    try:
        with csv_file.open(encoding="utf-8") as f:
            reader = csv.reader(f)

            try:
                header = next(reader)
            except StopIteration:
                raise ValueError("CSV без заголовка или пустой")

            if not header or all((h or "").strip() == "" for h in header):
                raise ValueError("CSV без заголовка")

            rows = [header] + [row for row in reader]
    except FileNotFoundError:
        raise

    if len(rows) <= 1:
        raise ValueError("Пустой CSV")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    max_widths = [max(len(str(cell)), 8) for cell in rows[0]]

    for row in rows:
        ws.append(row)

        if len(row) > len(max_widths):
            max_widths.extend([8] * (len(row) - len(max_widths)))

        for i, cell in enumerate(row):
            length = len(str(cell))
            if length > max_widths[i]:
                max_widths[i] = length

    for col_index, width in enumerate(max_widths, start=1):
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = max(width, 8)

    xlsx_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_file)
